import json
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.models.entities import (
    DocumentChunk,
    KnowledgeChunk,
    KnowledgeDocument,
    TopicSourceDocument,
)
from app.models.enums import DocumentStatus, DocumentType


def detect_doc_type(filename: str) -> DocumentType:
    suffix = Path(filename).suffix.lower()
    mapping = {
        ".pdf": DocumentType.pdf,
        ".docx": DocumentType.docx,
        ".xlsx": DocumentType.xlsx,
        ".txt": DocumentType.txt,
        ".md": DocumentType.markdown,
        ".markdown": DocumentType.markdown,
    }
    if suffix not in mapping:
        raise ValueError("Unsupported file extension")
    return mapping[suffix]


class IngestionService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def ingest_document(self, document: TopicSourceDocument) -> None:
        document.status = DocumentStatus.parsing
        self.db.add(document)
        self.db.commit()
        try:
            text_data = self._extract_text(document.file_path, document.doc_type)
            chunks = self._chunk_text(text_data)
            self.db.query(DocumentChunk).filter(
                DocumentChunk.document_id == document.id
            ).delete()
            for i, chunk in enumerate(chunks):
                row = DocumentChunk(
                    document_id=document.id,
                    topic_id=document.topic_id,
                    chunk_index=i,
                    text=chunk,
                    metadata_json=json.dumps(
                        {"filename": document.filename, "chunk_index": i}
                    ),
                )
                self.db.add(row)
                self.db.flush()
                self.db.execute(
                    text(
                        "INSERT INTO document_chunks_fts(chunk_id, topic_id, text) VALUES (:c, :t, :x)"
                    ),
                    {"c": row.id, "t": document.topic_id, "x": chunk},
                )
            document.status = DocumentStatus.indexed
            document.extraction_error = ""
            self.db.add(document)
            self.db.commit()
        except Exception as exc:
            document.status = DocumentStatus.failed
            document.extraction_error = str(exc)
            self.db.add(document)
            self.db.commit()

    def ingest_knowledge_document(self, document: KnowledgeDocument) -> None:
        document.status = DocumentStatus.parsing
        self.db.add(document)
        self.db.commit()
        try:
            text_data = self._extract_text(document.file_path, document.doc_type)
            chunks = self._chunk_text(text_data)
            self.db.query(KnowledgeChunk).filter(
                KnowledgeChunk.knowledge_document_id == document.id
            ).delete()
            self.db.execute(
                text(
                    "DELETE FROM knowledge_chunks_fts WHERE knowledge_document_id = :d"
                ),
                {"d": document.id},
            )
            for i, chunk in enumerate(chunks):
                row = KnowledgeChunk(
                    knowledge_document_id=document.id,
                    project_id=document.project_id,
                    chunk_index=i,
                    text=chunk,
                    metadata_json=json.dumps(
                        {
                            "filename": document.filename,
                            "chunk_index": i,
                            "project_id": document.project_id,
                        }
                    ),
                )
                self.db.add(row)
                self.db.flush()
                self.db.execute(
                    text(
                        """
                        INSERT INTO knowledge_chunks_fts(chunk_id, knowledge_document_id, project_id, text)
                        VALUES (:c, :d, :p, :x)
                        """
                    ),
                    {
                        "c": row.id,
                        "d": document.id,
                        "p": document.project_id,
                        "x": chunk,
                    },
                )

            document.status = DocumentStatus.indexed
            document.extraction_error = ""
            self.db.add(document)
            self.db.commit()
        except Exception as exc:
            document.status = DocumentStatus.failed
            document.extraction_error = str(exc)
            self.db.add(document)
            self.db.commit()

    def search_chunks(
        self, topic_id: str, query: str, limit: int = 8
    ) -> list[dict[str, str]]:
        rows = self.db.execute(
            text(
                """
                SELECT f.chunk_id, f.text
                FROM document_chunks_fts f
                WHERE f.topic_id = :topic_id AND f.text MATCH :q
                LIMIT :limit
                """
            ),
            {"topic_id": topic_id, "q": query, "limit": limit},
        )
        return [{"chunk_id": r[0], "text": r[1]} for r in rows.fetchall()]

    def _extract_text(self, path: str, doc_type: DocumentType) -> str:
        if doc_type == DocumentType.txt or doc_type == DocumentType.markdown:
            return Path(path).read_text(encoding="utf-8", errors="ignore")
        if doc_type == DocumentType.pdf:
            reader = PdfReader(path)
            return "\n".join([page.extract_text() or "" for page in reader.pages])
        if doc_type == DocumentType.docx:
            doc = Document(path)
            return "\n".join([p.text for p in doc.paragraphs])
        if doc_type == DocumentType.xlsx:
            wb = load_workbook(path, data_only=True)
            lines: list[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None]
                    if values:
                        lines.append(" | ".join(values))
            return "\n".join(lines)
        raise ValueError("Unsupported file type")

    def _chunk_text(self, text_data: str, max_chars: int = 1500) -> list[str]:
        words = text_data.split()
        chunks: list[str] = []
        current: list[str] = []
        length = 0
        for word in words:
            if length + len(word) + 1 > max_chars and current:
                chunks.append(" ".join(current))
                current = []
                length = 0
            current.append(word)
            length += len(word) + 1
        if current:
            chunks.append(" ".join(current))
        return chunks
